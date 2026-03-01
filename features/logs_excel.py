import asyncio
import os
import re
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import ContextTypes

from core.logging_ops import log_event
from core.session import end_session, touch_session, user_sessions
from core.time_utils import now_hk
from integrations.gmail import fetch_rthk_emails_for_excel
from ui.messages import SESSION_EXPIRED_TEXT


def parse_rthk_json_data(json_data: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """从JSON附件数据解析为Excel行数据"""
    rows: List[Dict[str, Any]] = []
    if not json_data or not isinstance(json_data, list):
        return rows

    for category in json_data:
        target = category.get("target", "")
        data_list = category.get("data", [])
        if not isinstance(data_list, list):
            continue

        for item in data_list:
            rows.append(
                {
                    "Post ID": item.get("post_id", ""),
                    "Target": target,
                    "Original Subject": item.get("title", ""),
                    "WP Title": item.get("wp_title", ""),
                    "Original URL": item.get("url", ""),
                    "Original Time Text": item.get("time_text", ""),
                    "新聞內文": item.get("body", ""),
                    "改寫後內文": item.get("wp_body", ""),
                }
            )

    return rows


def generate_rthk_excel(items: List[Dict[str, Any]]) -> tuple[str, int]:
    from openpyxl.cell.cell import Hyperlink

    os.makedirs("temp", exist_ok=True)
    filename = f"RTHK News改寫狀況一覽_{now_hk().strftime('%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join("temp", filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "RTHK Logs"

    headers = [
        "Post ID",
        "Target",
        "Original Subject",
        "WP Title",
        "Original URL",
        "Original Time Text",
        "新聞內文",
        "改寫後內文",
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据并设置格式
    for item in items:
        row_data = [item.get(h, "") for h in headers]
        ws.append(row_data)
        row_idx = ws.max_row

        # 设置 Original Subject 和 WP Title 的对齐（都贴上面，vertical="top"）
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")  # Original Subject
        ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")  # WP Title

        # 设置文本换行：Original URL, 新聞內文, 改寫後內文
        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")  # Original URL
        ws.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True, vertical="top")  # 新聞內文
        ws.cell(row=row_idx, column=8).alignment = Alignment(wrap_text=True, vertical="top")  # 改寫後內文

        # 设置 URL 为超链接
        url_value = row_data[4]  # Original URL 在第5列（索引4）
        if url_value and url_value.startswith("http"):
            url_cell = ws.cell(row=row_idx, column=5)
            url_cell.hyperlink = url_value
            url_cell.font = Font(color="0000FF", underline="single")

    # 设置列宽
    ws.column_dimensions["A"].width = 16  # Post ID
    ws.column_dimensions["B"].width = 14  # Target
    # Original Subject (C列) 自动调整宽度
    ws.column_dimensions["D"].width = 52  # WP Title
    ws.column_dimensions["E"].width = 52  # Original URL
    ws.column_dimensions["F"].width = 24  # Original Time Text
    ws.column_dimensions["G"].width = 60  # 新聞內文
    ws.column_dimensions["H"].width = 60  # 改寫後內文

    # 自动调整 Original Subject 列宽（根据内容）
    max_length = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            try:
                if cell.value:
                    # 计算中文字符长度（中文字符通常占2个字符宽度）
                    cell_length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_length = max(max_length, cell_length)
            except Exception:
                pass
    # 设置列宽，留一些边距
    ws.column_dimensions["C"].width = min(max(max_length + 2, 20), 100)

    wb.save(output_path)
    return output_path, len(items)


async def on_excel_export_rthk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    session_key = query.data.split("|")[1]

    if session_key not in user_sessions:
        await query.edit_message_text(SESSION_EXPIRED_TEXT)
        return

    touch_session(
        context=context,
        session_key=session_key,
        user_id=query.from_user.id,
        chat_id=query.message.chat.id,
        message_id=query.message.message_id,
    )
    session_data = user_sessions.get(session_key) or {}

    try:
        log_event(
            "logs_excel_rthk_click",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
        )
    except Exception:
        pass

    await query.edit_message_text("正在匯出 RTHK Logs（最近24小時，HKT）...\n請稍候。")

    output_path = ""
    try:
        log_event(
            "logs_excel_gmail_fetch_start",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "hours": 24},
        )

        emails = await asyncio.to_thread(fetch_rthk_emails_for_excel, 24, 500)
        log_event(
            "logs_excel_gmail_fetch_done",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "email_count": len(emails)},
        )

        all_items: List[Dict[str, Any]] = []
        for one in emails:
            json_data = one.get("json_data")
            all_items.extend(parse_rthk_json_data(json_data))

        log_event(
            "logs_excel_parse_done",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "item_count": len(all_items)},
        )

        output_path, row_count = await asyncio.to_thread(generate_rthk_excel, all_items)
        file_size = os.path.getsize(output_path)
        log_event(
            "logs_excel_generate_done",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "row_count": row_count, "path": output_path, "size": file_size},
        )

        with open(output_path, "rb") as fp:
            sent_message = await context.bot.send_document(
                chat_id=query.message.chat.id,
                document=fp,
                filename=os.path.basename(output_path),
                caption=f"RTHK Logs 導出（最近24小時 HKT），共 {row_count} 條。",
            )

        send_ok = bool(sent_message and getattr(sent_message, "document", None))
        log_event(
            "logs_excel_send_done",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "send_ok": send_ok, "chat_id": query.message.chat.id},
        )
        log_event(
            "logs_excel_send_verify",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "verified": send_ok},
        )

        if not send_ok:
            buttons = [[InlineKeyboardButton("⬅️ 返回", callback_data=f"logs_excel_export|{session_key}")]]
            await query.edit_message_text(
                "⚠️ Excel已生成，但發送驗證未通過，請重試。",
                reply_markup=InlineKeyboardMarkup(buttons),
            )
            return

        log_event(
            "logs_excel_end_session",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK"},
        )
        await end_session(
            application=context.application,
            session_key=session_key,
            reason_text="excel已生成，會話結束。",
            reason_code="logs_excel_export_done",
            user_id=query.from_user.id,
            chat_id=query.message.chat.id,
            message_id=query.message.message_id,
        )
    except Exception as e:
        log_event(
            "logs_excel_export_failed",
            session_key=session_key,
            session_id=session_data.get("session_id"),
            update=update,
            extra={"source": "RTHK", "error": str(e)},
        )
        buttons = [[InlineKeyboardButton("⬅️ 返回", callback_data=f"logs_excel_export|{session_key}")]]
        await query.edit_message_text(
            f"❌ Excel導出失敗：{e}",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
    finally:
        if output_path and os.path.exists(output_path):
            try:
                os.remove(output_path)
            except Exception:
                pass
